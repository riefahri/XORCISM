"""
import_threatevent.py — Imports the "Adversarial Threat Events" of the
Security-Scientist-NIST-800-30-Reference-Library.xlsx workbook into XTHREAT.THREATEVENT.
Jerome Athias - XORCISM

Sheet read: "Adversarial Threats".
The header is detected dynamically (an empty column A precedes the data):
  ID               → ReferentialID
  Kill-chain Phase → KCPhase
  Tier             → Tier (integer)
  Threat Event     → Description
VocabularyID is set with the "NIST-800-30" vocabulary (XORCISM.VOCABULARY).

Idempotent: get-or-create by ReferentialID.

Usage:
    python import_threatevent.py [--xlsx path.xlsx]
"""

import argparse
import os
import sys
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from sqlalchemy import event, text, Integer, BigInteger, SmallInteger
from sqlalchemy.orm import Mapper

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from xorcism_python.models.base import session_scope
from xorcism_python.utils import log

MODULE = "ImportThreatEvent"
SHEET = "Adversarial Threats"
VOCAB_NAME = "NIST-800-30"
DEFAULT_XLSX = os.path.join(
    os.path.dirname(__file__), "..", "..",
    "resources", "Security-Scientist-NIST-800-30-Reference-Library.xlsx",
)


# ─── Auto-increment of integer primary keys (cf. other importers) ───────

_pk_counters: dict = {}
_INT_TYPES = (Integer, BigInteger, SmallInteger)


def _auto_pk(mapper, connection, target) -> None:
    pk_cols = mapper.primary_key
    if len(pk_cols) != 1:
        return
    col = pk_cols[0]
    if not isinstance(col.type, _INT_TYPES):
        return
    attr = mapper.get_property_by_column(col).key
    if getattr(target, attr, None) is not None:
        return
    tbl = mapper.local_table.name
    key = (str(connection.engine.url), tbl)
    db_max = connection.execute(
        text(f'SELECT COALESCE(MAX("{col.name}"), 0) FROM "{tbl}"')
    ).scalar()
    nxt = max(_pk_counters.get(key, 0), int(db_max or 0)) + 1
    _pk_counters[key] = nxt
    setattr(target, attr, nxt)


event.listen(Mapper, "before_insert", _auto_pk)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).replace("\r\n", " ").replace("\n", " ").replace("\t", " ").strip()
    while "  " in s:
        s = s.replace("  ", " ")
    return s or None


def _to_int(v) -> Optional[int]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def _setup_vocabulary(xorcism_session) -> int:
    from xorcism_python.models.xorcism import VOCABULARY
    vocab = xorcism_session.query(VOCABULARY).filter_by(VocabularyName=VOCAB_NAME).first()
    if vocab is None:
        vocab = VOCABULARY(VocabularyName=VOCAB_NAME, CreatedDate=_now())
        xorcism_session.add(vocab)
        xorcism_session.flush()
        log(MODULE, f"VOCABULARY {VOCAB_NAME} créé (ID={vocab.VocabularyID})")
    return vocab.VocabularyID


def _find_header(rows, required):
    """Finds the header row containing all the `required` columns.
    Returns (idx, cols) where cols maps column_name -> index."""
    for i, row in enumerate(rows):
        vals = [(_clean(c) or "") for c in row]
        if all(r in vals for r in required):
            cols = {}
            for j, v in enumerate(vals):
                if v and v not in cols:
                    cols[v] = j
            return i, cols
    raise ValueError(f"Entête introuvable (colonnes {' / '.join(required)})")


# ─── Import ────────────────────────────────────────────────────────────────────

def parse_xlsx(xlsx_path: str) -> None:
    log(MODULE, f"Lecture {xlsx_path} — onglet « {SHEET} »")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"Onglet « {SHEET} » absent. Onglets : {wb.sheetnames}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    hdr_idx, cols = _find_header(rows, ["ID", "Kill-chain Phase", "Threat Event"])
    ci_id = cols["ID"]
    ci_kc = cols["Kill-chain Phase"]
    ci_te = cols["Threat Event"]
    ci_tier = cols.get("Tier")
    log(MODULE, f"Entête ligne {hdr_idx + 1} ; colonnes ID={ci_id} KC={ci_kc} Tier={ci_tier} TE={ci_te}")

    with session_scope("XORCISM") as xorcism_session:
        vocab_id = _setup_vocabulary(xorcism_session)
        xorcism_session.commit()

    from xorcism_python.models.xthreat import THREATEVENT
    created = updated = 0
    with session_scope("XTHREAT") as session:
        for row in rows[hdr_idx + 1:]:
            ref = _clean(row[ci_id]) if ci_id < len(row) else None
            if not ref:
                continue  # empty row / without an identifier
            kc = _clean(row[ci_kc]) if ci_kc < len(row) else None
            desc = _clean(row[ci_te]) if ci_te < len(row) else None
            tier = _to_int(row[ci_tier]) if (ci_tier is not None and ci_tier < len(row)) else None

            te = session.query(THREATEVENT).filter_by(ReferentialID=ref).first()
            if te is None:
                te = THREATEVENT(ReferentialID=ref, VocabularyID=vocab_id)
                session.add(te)
                created += 1
            else:
                updated += 1
            te.KCPhase = kc
            te.Tier = tier
            te.Description = desc
            te.VocabularyID = vocab_id
            session.flush()

            if (created + updated) % 50 == 0:
                session.commit()
                log(MODULE, f"  {created + updated} lignes traitées…")
        session.commit()

    log(MODULE, f"Import terminé : {created} créées, {updated} mises à jour.")


SHEET_NON_ADV = "Non-Adversarial Threats"
VOCAB_ID_NON_ADV = 5  # NIST-800-30 (imposed)


def parse_non_adversarial(xlsx_path: str) -> None:
    """Imports the "Non-Adversarial Threats" sheet.
    Mapping: ID(B)→ReferentialID, Category(D)→Category, Tier(E)→Tier,
    Threat Event(F)→Description. VocabularyID = 5. KCPhase left empty.
    Idempotent (get-or-create by ReferentialID)."""
    log(MODULE, f"Lecture {xlsx_path} — onglet « {SHEET_NON_ADV} »")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NON_ADV not in wb.sheetnames:
        raise ValueError(f"Onglet « {SHEET_NON_ADV} » absent. Onglets : {wb.sheetnames}")
    ws = wb[SHEET_NON_ADV]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    hdr_idx, cols = _find_header(rows, ["ID", "Category", "Tier", "Threat Event"])
    ci_id = cols["ID"]
    ci_cat = cols["Category"]
    ci_tier = cols["Tier"]
    ci_te = cols["Threat Event"]
    log(MODULE, f"Entête ligne {hdr_idx + 1} ; colonnes ID={ci_id} Category={ci_cat} Tier={ci_tier} TE={ci_te}")

    from xorcism_python.models.xthreat import THREATEVENT
    created = updated = 0
    with session_scope("XTHREAT") as session:
        for row in rows[hdr_idx + 1:]:
            ref = _clean(row[ci_id]) if ci_id < len(row) else None
            if not ref:
                continue
            cat = _clean(row[ci_cat]) if ci_cat < len(row) else None
            tier = _to_int(row[ci_tier]) if ci_tier < len(row) else None
            desc = _clean(row[ci_te]) if ci_te < len(row) else None

            te = session.query(THREATEVENT).filter_by(ReferentialID=ref).first()
            if te is None:
                te = THREATEVENT(ReferentialID=ref, VocabularyID=VOCAB_ID_NON_ADV)
                session.add(te)
                created += 1
            else:
                updated += 1
            te.Category = cat
            te.Tier = tier
            te.Description = desc
            te.VocabularyID = VOCAB_ID_NON_ADV
            session.flush()

            if (created + updated) % 50 == 0:
                session.commit()
                log(MODULE, f"  {created + updated} lignes traitées…")
        session.commit()

    log(MODULE, f"Import non-adversarial terminé : {created} créées, {updated} mises à jour.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Threat Events (NIST 800-30) dans XTHREAT.THREATEVENT")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Chemin du classeur xlsx")
    parser.add_argument(
        "--sheet", choices=["adversarial", "non-adversarial", "both"], default="both",
        help="Onglet(s) à importer (défaut : both)"
    )
    args = parser.parse_args()
    if not os.path.exists(args.xlsx):
        raise SystemExit(f"Fichier introuvable : {args.xlsx}")
    if args.sheet in ("adversarial", "both"):
        parse_xlsx(args.xlsx)
    if args.sheet in ("non-adversarial", "both"):
        parse_non_adversarial(args.xlsx)


if __name__ == "__main__":
    main()
