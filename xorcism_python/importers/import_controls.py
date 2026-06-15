"""
import_controls.py — Imports the "Controls & Mitigations" sheet of the
Security-Scientist-NIST-800-30-Reference-Library.xlsx workbook into XORCISM.CONTROL.
Jerome Athias - XORCISM

Header (row 4): Control Theme | NIST 800-53 | ISO 27002 | CIS v8 | Minimal |
Balanced | Comprehensive. Data from row 5.
Mapping (by column letter):
  B → ControlName
  C → NIST
  D → ISO
  E → CIS
  F → Minimal
  G → Balanced
  H → Comprehensive
VocabularyID = 1 (XORCISM).

Idempotent: get-or-update by ControlName.

Usage:
    python import_controls.py [--xlsx path.xlsx]
"""

import argparse
import os
import sys
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from sqlalchemy import event, text, Integer, BigInteger, SmallInteger
from sqlalchemy.orm import Mapper

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from xorcism_python.models.base import session_scope
from xorcism_python.utils import log

MODULE = "ImportControls"
SHEET = "Controls & Mitigations"
VOCAB_ID = 1  # XORCISM
DEFAULT_XLSX = os.path.join(
    os.path.dirname(__file__), "..", "..",
    "resources", "Security-Scientist-NIST-800-30-Reference-Library.xlsx",
)

# Column indices (0-based); column A is empty → B=1 … H=7.
CI_NAME = 1          # B  Control Theme
CI_NIST = 2          # C  NIST 800-53
CI_ISO = 3           # D  ISO 27002
CI_CIS = 4           # E  CIS v8
CI_MINIMAL = 5       # F  Minimal
CI_BALANCED = 6      # G  Balanced
CI_COMPREHENSIVE = 7  # H  Comprehensive


# ─── Auto-increment of integer primary keys (cf. other importers) ───────

_pk_counters: dict = {}
_INT_TYPES = (Integer, BigInteger, SmallInteger)


def _auto_pk(mapper, connection, target) -> None:
    pk_cols = mapper.primary_key
    if len(pk_cols) != 1:
        return
    col = pk_cols[0]
    if not isinstance(col.type, _INT_TYPES):
        return
    attr = mapper.get_property_by_column(col).key
    if getattr(target, attr, None) is not None:
        return
    tbl = mapper.local_table.name
    key = (str(connection.engine.url), tbl)
    db_max = connection.execute(
        text(f'SELECT COALESCE(MAX("{col.name}"), 0) FROM "{tbl}"')
    ).scalar()
    nxt = max(_pk_counters.get(key, 0), int(db_max or 0)) + 1
    _pk_counters[key] = nxt
    setattr(target, attr, nxt)


event.listen(Mapper, "before_insert", _auto_pk)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).replace("\r\n", " ").replace("\n", " ").replace("\t", " ").strip()
    while "  " in s:
        s = s.replace("  ", " ")
    return s or None


def _cell(row, idx) -> Optional[str]:
    return _clean(row[idx]) if idx < len(row) else None


def _find_header(rows) -> int:
    """Index of the header row (containing "Control Theme" and "Comprehensive")."""
    for i, row in enumerate(rows):
        vals = [(_clean(c) or "") for c in row]
        if "Control Theme" in vals and "Comprehensive" in vals:
            return i
    raise ValueError("Entête introuvable (Control Theme / Comprehensive)")


# ─── Import ────────────────────────────────────────────────────────────────────

def parse_xlsx(xlsx_path: str) -> None:
    log(MODULE, f"Lecture {xlsx_path} — onglet « {SHEET} »")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"Onglet « {SHEET} » absent. Onglets : {wb.sheetnames}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    hdr_idx = _find_header(rows)
    log(MODULE, f"Entête ligne {hdr_idx + 1} ; données à partir de la ligne {hdr_idx + 2}")

    from xorcism_python.models.xorcism import CONTROL
    created = updated = 0
    with session_scope("XORCISM") as session:
        for row in rows[hdr_idx + 1:]:
            name = _cell(row, CI_NAME)
            if not name:
                continue
            ctrl = session.query(CONTROL).filter_by(ControlName=name).first()
            if ctrl is None:
                ctrl = CONTROL(ControlName=name, CreatedDate=_now())
                session.add(ctrl)
                created += 1
            else:
                updated += 1
            ctrl.NIST = _cell(row, CI_NIST)
            ctrl.ISO = _cell(row, CI_ISO)
            ctrl.CIS = _cell(row, CI_CIS)
            ctrl.Minimal = _cell(row, CI_MINIMAL)
            ctrl.Balanced = _cell(row, CI_BALANCED)
            ctrl.Comprehensive = _cell(row, CI_COMPREHENSIVE)
            ctrl.VocabularyID = VOCAB_ID
            session.flush()
        session.commit()

    log(MODULE, f"Import terminé : {created} créés, {updated} mis à jour.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Controls & Mitigations dans XORCISM.CONTROL")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Chemin du classeur xlsx")
    args = parser.parse_args()
    if not os.path.exists(args.xlsx):
        raise SystemExit(f"Fichier introuvable : {args.xlsx}")
    parse_xlsx(args.xlsx)


if __name__ == "__main__":
    main()
