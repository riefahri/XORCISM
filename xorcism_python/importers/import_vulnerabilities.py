"""
import_vulnerabilities.py — Imports the "Vulnerabilities" sheet of the
Security-Scientist-NIST-800-30-Reference-Library.xlsx workbook into XVULNERABILITY.VULNERABILITY.
Jerome Athias - XORCISM

Header (row 4): ID | Domain | Tier | Type | Kind | Weakness / Condition.
Data from row 5. Mapping (by column letter):
  B (ID)     → VULGUID  (and copied into VULReferentialID)
  C (Domain) → VULDomain
  D (Tier)   → Tier (integer)
  E (Type)   → VULType
  G          → VULDescription   ("Weakness / Condition")
VocabularyID = 5 (NIST-800-30).

Idempotent: get-or-update by VULGUID (batched lookup in one query,
because the table is large).

Usage:
    python import_vulnerabilities.py [--xlsx path.xlsx]
"""

import argparse
import os
import sys
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from sqlalchemy import event, text, Integer, BigInteger, SmallInteger
from sqlalchemy.orm import Mapper

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from xorcism_python.models.base import session_scope
from xorcism_python.utils import log

MODULE = "ImportVulnerabilities"
SHEET = "Vulnerabilities"
VOCAB_ID = 5  # NIST-800-30
DEFAULT_XLSX = os.path.join(
    os.path.dirname(__file__), "..", "..",
    "resources", "Security-Scientist-NIST-800-30-Reference-Library.xlsx",
)

# Column indices (0-based); column A is empty → B=1, C=2, …, G=6.
CI_ID = 1     # B
CI_DOMAIN = 2  # C
CI_TIER = 3    # D
CI_TYPE = 4    # E
CI_DESC = 6    # G


# ─── Auto-increment of integer primary keys (cf. other importers) ───────

_pk_counters: dict = {}
_INT_TYPES = (Integer, BigInteger, SmallInteger)


def _auto_pk(mapper, connection, target) -> None:
    pk_cols = mapper.primary_key
    if len(pk_cols) != 1:
        return
    col = pk_cols[0]
    if not isinstance(col.type, _INT_TYPES):
        return
    attr = mapper.get_property_by_column(col).key
    if getattr(target, attr, None) is not None:
        return
    tbl = mapper.local_table.name
    key = (str(connection.engine.url), tbl)
    db_max = connection.execute(
        text(f'SELECT COALESCE(MAX("{col.name}"), 0) FROM "{tbl}"')
    ).scalar()
    nxt = max(_pk_counters.get(key, 0), int(db_max or 0)) + 1
    _pk_counters[key] = nxt
    setattr(target, attr, nxt)


event.listen(Mapper, "before_insert", _auto_pk)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).replace("\r\n", " ").replace("\n", " ").replace("\t", " ").strip()
    while "  " in s:
        s = s.replace("  ", " ")
    return s or None


def _to_int(v) -> Optional[int]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def _find_header(rows) -> int:
    """Index of the header row (containing ID, Domain, Tier, Type)."""
    for i, row in enumerate(rows):
        vals = [(_clean(c) or "") for c in row]
        if "ID" in vals and "Domain" in vals and "Tier" in vals and "Type" in vals:
            return i
    raise ValueError("Entête introuvable (ID / Domain / Tier / Type)")


# ─── Import ────────────────────────────────────────────────────────────────────

def parse_xlsx(xlsx_path: str) -> None:
    log(MODULE, f"Lecture {xlsx_path} — onglet « {SHEET} »")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"Onglet « {SHEET} » absent. Onglets : {wb.sheetnames}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    hdr_idx = _find_header(rows)
    log(MODULE, f"Entête ligne {hdr_idx + 1} ; données à partir de la ligne {hdr_idx + 2}")

    # Builds the list of records (deduplicated by GUID)
    records = []
    seen = set()
    for row in rows[hdr_idx + 1:]:
        guid = _clean(row[CI_ID]) if CI_ID < len(row) else None
        if not guid or guid in seen:
            continue
        seen.add(guid)
        records.append({
            "guid": guid,
            "domain": _clean(row[CI_DOMAIN]) if CI_DOMAIN < len(row) else None,
            "tier": _to_int(row[CI_TIER]) if CI_TIER < len(row) else None,
            "type": _clean(row[CI_TYPE]) if CI_TYPE < len(row) else None,
            "desc": _clean(row[CI_DESC]) if CI_DESC < len(row) else None,
        })
    log(MODULE, f"{len(records)} lignes distinctes trouvées")
    if not records:
        return

    from xorcism_python.models.xvulnerability import VULNERABILITY
    created = updated = 0
    with session_scope("XVULNERABILITY") as session:
        # Fetches the existing VULNERABILITY rows in one query (by GUID) → idempotence
        guids = [r["guid"] for r in records]
        existing = {}
        CHUNK = 500
        for k in range(0, len(guids), CHUNK):
            sub = guids[k:k + CHUNK]
            for vid, g in session.query(VULNERABILITY.VulnerabilityID, VULNERABILITY.VULGUID).filter(
                VULNERABILITY.VULGUID.in_(sub)
            ).all():
                existing[g] = vid

        for r in records:
            vobj = None
            if r["guid"] in existing:
                vobj = session.get(VULNERABILITY, existing[r["guid"]])
            if vobj is None:
                vobj = VULNERABILITY(VULGUID=r["guid"], CreatedDate=_now())
                session.add(vobj)
                created += 1
            else:
                updated += 1
            vobj.VULReferentialID = r["guid"]   # same value as VULGUID
            vobj.VULDomain = r["domain"]
            vobj.Tier = r["tier"]
            vobj.VULType = r["type"]
            vobj.VULDescription = r["desc"]
            vobj.VocabularyID = VOCAB_ID
            session.flush()

            if (created + updated) % 50 == 0:
                session.commit()
                log(MODULE, f"  {created + updated} lignes traitées…")
        session.commit()

    log(MODULE, f"Import terminé : {created} créées, {updated} mises à jour.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Vulnerabilities (NIST 800-30) dans XVULNERABILITY.VULNERABILITY")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Chemin du classeur xlsx")
    args = parser.parse_args()
    if not os.path.exists(args.xlsx):
        raise SystemExit(f"Fichier introuvable : {args.xlsx}")
    parse_xlsx(args.xlsx)


if __name__ == "__main__":
    main()
