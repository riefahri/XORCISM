"""
import_caiq.py — Import a CSA CAIQ workbook into XCOMPLIANCE as a QUESTIONNAIRE.

The CAIQ (Consensus Assessments Initiative Questionnaire) and the CCM it maps to are
Cloud Security Alliance copyrighted material: download your own copy from CSA; this
importer only loads the file you provide into your own instance.

Each CAIQ row → a QUESTION (QuestionName = Question ID e.g. 'A&A-01.1';
QuestionText = the question; QuestionDescription = the CCM domain code), linked to a
new QUESTIONNAIRE through QUESTIONFORQUESTIONNAIRE. Mirrors the app's Excel import
(importQuestionnaireFromExcel). Idempotent by questionnaire name (--force to re-import).

Usage:
    python import_caiq.py --file <CAIQ.xlsx> [--name "CAIQ v4.0.3 (CSA STAR)"] [--force]
"""
from __future__ import annotations

import argparse
import os
import sqlite3
import sys
from datetime import datetime, timezone
from uuid import uuid4

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from xorcism_python import config  # noqa: E402

DB_PATH = os.path.join(config.DB_DIR, "XCOMPLIANCE.db")


def now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def log(msg: str) -> None:
    print(f"[ImportCAIQ] {msg}", flush=True)


def parse_caiq(path: str):
    """Return (sheet_name, [{name, text, domain}]) from a CAIQ workbook."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((n for n in wb.sheetnames if n.lower().startswith("caiq")), None)
    if sheet is None:
        others = [n for n in wb.sheetnames if n.lower() != "introduction"]
        sheet = others[0] if others else wb.sheetnames[0]
    rows = list(wb[sheet].iter_rows(values_only=True))

    hdr = qid_col = qtext_col = None
    for i, r in enumerate(rows):
        cells = [("" if c is None else str(c)).strip().lower() for c in r]
        if "question id" in cells:
            hdr = i
            qid_col = cells.index("question id")
            qtext_col = cells.index("question") if "question" in cells else qid_col + 1
            break
    if hdr is None:
        raise RuntimeError(f"no 'Question ID' header found in sheet '{sheet}'")

    out = []
    for r in rows[hdr + 1:]:
        qid = ("" if r[qid_col] is None else str(r[qid_col])).strip()
        qtext = ("" if qtext_col >= len(r) or r[qtext_col] is None else str(r[qtext_col])).strip()
        if not qid:
            continue
        out.append({"name": qid, "text": qtext or qid, "domain": qid.split("-", 1)[0] if "-" in qid else ""})
    return sheet, out


def ensure_columns(cur: sqlite3.Cursor) -> None:
    """Add the columns the app's Excel-import uses, if a base schema lacks them."""
    def add(table: str, cols: dict) -> None:
        existing = {r[1] for r in cur.execute(f'PRAGMA table_info("{table}")').fetchall()}
        for name, typ in cols.items():
            if name not in existing:
                cur.execute(f'ALTER TABLE "{table}" ADD COLUMN "{name}" {typ}')
    add("QUESTIONNAIRE", {"FileName": "TEXT", "TenantID": "INTEGER"})
    add("QUESTION", {"QuestionText": "TEXT", "QuestionType": "TEXT", "DefaultAnswer": "TEXT", "TenantID": "INTEGER"})
    add("QUESTIONFORQUESTIONNAIRE", {"DisplayOrder": "INTEGER", "TenantID": "INTEGER"})


def main() -> None:
    ap = argparse.ArgumentParser(description="Import a CSA CAIQ workbook into XCOMPLIANCE")
    ap.add_argument("--file", required=True, help="Path to the CAIQ .xlsx (downloaded from CSA)")
    ap.add_argument("--name", help="Questionnaire name (default derived from the sheet)")
    ap.add_argument("--force", action="store_true", help="Import even if a questionnaire with that name exists")
    _env_t = (os.getenv("XORCISM_IMPORT_TENANT_ID") or "").strip()
    ap.add_argument("--tenant", type=int, default=(int(_env_t) if _env_t.lstrip("-").isdigit() else None),
                    help="TenantID to assign (QUESTIONNAIRE/QUESTION are tenant-scoped). Default: $XORCISM_IMPORT_TENANT_ID or NULL.")
    args = ap.parse_args()

    sheet, questions = parse_caiq(args.file)
    if not questions:
        log("No questions parsed — aborting.")
        return
    qname = args.name or f"{sheet.replace('CAIQv', 'CAIQ v')} (CSA STAR)"
    log(f"{len(questions)} questions parsed from sheet '{sheet}'")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA busy_timeout=15000")
    cur = conn.cursor()
    ensure_columns(cur)

    existing = cur.execute("SELECT QuestionnaireID FROM QUESTIONNAIRE WHERE QuestionnaireName=?", (qname,)).fetchone()
    if existing and not args.force:
        log(f"A questionnaire named '{qname}' already exists (ID {existing[0]}). Use --force to import again.")
        conn.close()
        return

    qid = cur.execute("SELECT COALESCE(MAX(QuestionnaireID),0)+1 FROM QUESTIONNAIRE").fetchone()[0]
    cur.execute(
        "INSERT INTO QUESTIONNAIRE (QuestionnaireID, QuestionnaireName, QuestionnaireDescription, FileName, CreatedDate, TenantID) VALUES (?,?,?,?,?,?)",
        (qid, qname[:300], "Cloud Security Alliance CAIQ (CSA STAR Security Questionnaire)", os.path.basename(args.file)[:300], now(), args.tenant),
    )
    maxq = cur.execute("SELECT COALESCE(MAX(QuestionID),0) FROM QUESTION").fetchone()[0]
    maxl = cur.execute("SELECT COALESCE(MAX(QuestionForQuestionnaireID),0) FROM QUESTIONFORQUESTIONNAIRE").fetchone()[0]
    for order, q in enumerate(questions):
        maxq += 1
        cur.execute(
            "INSERT INTO QUESTION (QuestionID, QuestionGUID, QuestionName, QuestionText, QuestionDescription, QuestionType, DefaultAnswer, CreatedDate, TenantID) VALUES (?,?,?,?,?,?,?,?,?)",
            (maxq, str(uuid4()), q["name"][:500], q["text"], q["domain"], "boolean", "", now(), args.tenant),
        )
        maxl += 1
        cur.execute(
            "INSERT INTO QUESTIONFORQUESTIONNAIRE (QuestionForQuestionnaireID, QuestionnaireID, QuestionID, DisplayOrder, CreatedDate, TenantID) VALUES (?,?,?,?,?,?)",
            (maxl, qid, maxq, order, now(), args.tenant),
        )
    conn.commit()
    conn.close()
    log(f"Done ({now()}): QUESTIONNAIRE #{qid} '{qname}' + {len(questions)} questions imported.")


if __name__ == "__main__":
    main()
